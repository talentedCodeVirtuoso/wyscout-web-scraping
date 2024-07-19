import os
import sys
import openpyxl
from time import sleep
import pandas as pd
from bs4 import BeautifulSoup
from datetime import datetime

import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager


write_frequency = 2
sleep_timer = 8

################################ selenium functions ###############################################
options = Options()
options.add_argument("--window-size=1920,1080")
options.add_argument("--disable-extensions")
options.add_argument("--start-maximized")
# options.add_argument('--headless')
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")


service = Service(ChromeDriverManager().install())


## login to platform
def login():
    url = "https://platform.wyscout.com/app/?/"
    username = "jamie.davies02@hotmail.co.uk"
    password = "wyscout2"
    try:
        print("Logging in", end="\r")
        driver = webdriver.Chrome(options=options, service=service)
        driver.get(url)
        driver.maximize_window()
        WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.ID, "login-page"))
        )
        sleep(1)
        driver.find_element(
            By.CSS_SELECTOR, "input[type='email'][id='email']"
        ).send_keys(username)
        sleep(2)
        driver.find_element(
            By.CSS_SELECTOR, "input[type='password'][id='password']"
        ).send_keys(password)
        sleep(1)
        driver.find_element(
            By.CSS_SELECTOR, "button[type='submit'][id='logIn']"
        ).click()
        print("Logged in. . .")
        force_login_container = "login_restrictor_dialog_container"
        force_login_window = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, force_login_container))
        )
        if force_login_window:
            sleep(2)
            driver.find_element(By.CLASS_NAME, "btn2_zFM").click()
        print("-" * 70)
    except selenium.common.exceptions.TimeoutException as err:
        pass
    except Exception as err:
        print(f"Exception while logging in.", err)
        print("x" * 70)
    return driver


## function to call the mouse clicks
def click_items(driver, x):
    print("click buttons", end="\r")
    index, country, league, conference, team, _ = x

    driver, _ = select_item(driver, country, "detail_0_home_navy")
    if not _:
        return driver, None

    driver, leagues = select_item(driver, league, "detail_0_area_navy")
    if not leagues:
        return driver, None

    if conference:
        driver, conference = select_item(
            driver, conference, "detail_0_competition_navy_0"
        )
        if not conference:
            return driver, None

    driver, teams = select_item(driver, team, "detail_0_competition_navy")
    if not teams:
        return driver, None

    driver, stats = select_stats(driver)
    if not stats:
        update_cell([index], "NoStats")
        return driver, None
    driver = select_display(driver)
    driver = select_home(driver)
    return driver, "ok"


## select country / league / team
def select_item(driver, item, elem_id):
    try:
        WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.ID, elem_id))
        )
        print("selecting", item, end="\r")
        sleep(sleep_timer)
        elements = driver.find_elements(By.CLASS_NAME, "gears-list-item")
        for element in elements:
            if element.text == item:
                print(f"{len(elements):>3} CLICKED {element.text}")
                element.click()
                break
    except Exception as Err:
        print(f"Exception while selecting {item}")
        return driver, None
    return driver, elements


## click stats tab
def select_stats(driver):
    print("Selecting Stats", end="\r")
    try:
        WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.ID, "detail_0_team_tab_stats"))
        ).click()
        print("--- CLICKED Stats Tab")
    except Exception as Err:
        print("Exception while selecting Stats")
        print(
            f"XXX- No Stats for {index} {country} {team} {league} ({start} - {stop}) -XXX"
        )
        print("-" * 70)
        return driver, None
    return driver, True


## select Main from custom Display
def select_display(driver):
    try:
        WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.CLASS_NAME, "Select-control"))
        )
        print("Selecting Main View", end="\r")
        sleep(sleep_timer)
        select = driver.find_elements(By.CLASS_NAME, "Select")[-1]
        select.click()  # open select menu
        items = select.find_elements(
            By.CSS_SELECTOR, "div[class^='Preset__level-option']"
        )
        for x in items:
            if x.text.strip() == "Main":
                x.click()  # click 'Main' from DISPLAY menu
                print("--- CLICKED Main Display")
                break
        select.click()  # hide select menu
    except Exception as Err:
        print("Exception while selecting Main filter")
    return driver


## deselect Away teams
def select_home(driver):
    try:
        WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "table"))
        )
        print("Selecting Home teams", end="\r")
        sleep(sleep_timer)
        away_filter = driver.find_elements(
            By.CSS_SELECTOR, "div[class^='teamstats__Buttons-module__wrapper']"
        )[-1]
        away_filter.click()
        items = away_filter.find_elements(
            By.CSS_SELECTOR, "div[class^='teamstats__FilterOpened-module__group-item']"
        )
        for x in items:
            if x.text.strip() == "Away":
                x.click()
                print("--- CLICKED Away filter")
                break
        away_filter.click()
    except Exception as Err:
        print("Exception while selecting Away filter")
    return driver


## logout
def logout(driver):
    try:
        WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.ID, "account_button"))
        ).click()
        WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.ID, "user_actions"))
        ).click()
        tmp = driver.find_element(By.ID, "user_actions")
        btns = tmp.find_elements(By.CLASS_NAME, "gears-button")
        for btn in btns:
            if not btn.text == "Logout":
                continue
            btn.click()
            print("Logged out of Wyscout portal.")
            sleep(2)
            print("Closing the browser!")
            driver.quit()
            return
    except Exception as Err:
        print("Exception while logging out")


## scrol down the table
def scroll_down_table(driver):
    i = 0
    sleep(2)
    while True:
        driver.find_element(By.CSS_SELECTOR, "body").send_keys(Keys.DOWN)
        i += 1
        print("Scrolling down", i, end="\r")
        if i > 99:
            break
    sleep(4)
    return driver


################################ parse the data ###############################################

col_headers = [
    "league_name",
    "match_date",
    "home_team",
    "home_goal",
    "away_team",
    "away_goal",
    "home_xG",
    "away_xG",
    "home_shots",
    "home_ontarget",
    "away_shots",
    "away_ontarget",
    "home_passes",
    "home_completed_passes",
    "away_passes",
    "away_completed_passes",
    "home_corners",
    "home_corners_ shots",
    "away_corners",
    "away_corners_shots",
    "home_penalties",
    "home_converted",
    "away_penalties",
    "away_converted",
    "home_touches_in_box",
    "away_touches_in_box",
    "home_offside",
    "away_offside",
    "home_fouls",
    "away_fouls",
    "home_yellow_cards",
    "away_yellow_cards",
    "home_red_cards",
    "away_red_cards",
    "home_throw_ins",
    "home_accurate",
    "away_throw_ins",
    "away_accurate",
    "home_goal_kicks",
    "away_goal_kicks",
]

## lambda functions
pen = lambda x: x.split("/")[0] if "/" in x.split()[0] else x.split()[0]
con = lambda x: x.split("/")[-1][:1] if "/" in x.split()[0] else x.split()[0]
dates = lambda x: x.split()[-1]
convert = lambda x: datetime.strptime(x, "%d.%m.%Y")


## find table, scroll down, create soup & return table
def get_soup(index, driver, err_count=0):
    print("--- Parsing data table")
    sleep(sleep_timer)
    WebDriverWait(driver, 120).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "table"))
    )
    driver.find_elements(By.CSS_SELECTOR, "td")[0].click()
    scroll_down_table(driver)
    html = driver.page_source
    try:
        soup = BeautifulSoup(html, "lxml")
        table = soup.find("table")
    except:
        if err_count > 4:
            print(
                f"XXX No Table Data for {index} {country} {team} {league} {start} {stop} XXX"
            )
            print("-" * 70)
            update_cell([index], "NoTable")
            return None
        print(i, "Soup not created")
        get_soup(driver, err_count)
    return table


## create json dict for each row
def create_js(h):
    return {
        col_headers[0]: " ".join(h[1].split()[:-1]).rstrip(","),  # league_name
        col_headers[1]: dates(h[1]),  # match_date
        col_headers[2]: h[0].split(":")[0][:-1].strip(),  # home_team
        col_headers[3]: h[2],  # home_goal
        col_headers[4]: h[0].split(":")[-1][1:].strip(),  # away_team
        col_headers[5]: h[15],  # away_goal
        col_headers[6]: h[3],  # home_xG
        col_headers[7]: h[16],  # away_XG
        col_headers[8]: h[4].split()[0].split("/")[0],  # home_shots
        col_headers[9]: h[4].split()[0].split("/")[-1],  # home_ontarget
        col_headers[10]: h[17].split()[0].split("/")[0],  # away_shots
        col_headers[11]: h[17].split()[0].split("/")[-1],  # away_ontarget
        col_headers[12]: h[5].split()[0].split("/")[0],  # home_passes
        col_headers[13]: h[5].split()[0].split("/")[-1],  # home_completed_passes
        col_headers[14]: h[18].split()[0].split("/")[0],  # away_passes
        col_headers[15]: h[18].split()[0].split("/")[-1],  # away_completed_passes
        col_headers[16]: h[6].split()[0].split("/")[0],  # home_corners
        col_headers[17]: h[6].split()[0].split("/")[-1],  # home_cornsers_shots
        col_headers[18]: h[19].split()[0].split("/")[0],  # away_corners
        col_headers[19]: h[19].split()[0].split("/")[-1],  # away_corners_shots
        col_headers[20]: pen(h[7]),  # home_penalties
        col_headers[21]: con(h[7]),  # home_converted
        col_headers[22]: pen(h[20]),  # away_penalties
        col_headers[23]: con(h[20]),  # away_converted
        col_headers[24]: h[8].split()[0],  # home_touches_in_box
        col_headers[25]: h[21].split()[0],  # away_touches_in_box
        col_headers[26]: h[9].split()[0],  # home_offside
        col_headers[27]: h[22].split()[0],  # away_offside
        col_headers[28]: h[10].split()[0],  # home_fouls
        col_headers[29]: h[23].split()[0],  # away_fouls
        col_headers[30]: h[11].split()[0],  # home_yellow_cards
        col_headers[31]: h[24].split()[0],  # away_yellow_cards
        col_headers[32]: h[12].split()[0],  # home_red_cards
        col_headers[33]: h[25].split()[0],  # away_red_cards
        col_headers[34]: h[13].split()[0].split("/")[0],  # home_throw_ins
        col_headers[35]: h[13].split()[0].split("/")[-1],  # home_accurate
        col_headers[36]: h[26].split()[0].split("/")[0],  # away_throw_ins
        col_headers[37]: h[26].split()[0].split("/")[-1],  # away_accutate
        col_headers[38]: h[14].split()[0],  # home_goal_kicsk
        col_headers[39]: h[27].split()[0],  # away_goal_kicks
    }


## join away & home rows
def join_rows(rows):
    joined_list = []
    for i in range(len(rows)):
        if i % 2 != 0:
            continue
        rows[i].extend(rows[i + 1])
        joined_list.append(rows[i])
    return joined_list


## filter dates outside start and stop
def filter_dates(start, stop, inp):
    return convert(start) <= convert(inp) <= convert(stop)


## extract data from the table
def xtract_data(start, stop, table):
    rows = [tr for tr in table.find_all("tr")][3:]  # remove header rows
    joined_rows = join_rows(rows)  # join home & away rows

    data = []
    for row in joined_rows:
        cells = [y.text for y in row.find_all("td")][1:]  # find all cells
        col_1 = [y.text for y in row.find_all("div") if y.text]  # find team/league/date
        col_1.extend(cells)  # combine team with row cells
        if not filter_dates(start, stop, dates(col_1[1])):
            continue  # remove dates not in range
        js = create_js(col_1)
        data.append(js)  # create list of json
    return data


############################################ Excel file operations ##################################################
## def write data
def write_excel(df, out_file="wyscout.xlsx"):
    if not os.path.exists(out_file):
        tmp_df = pd.DataFrame(columns=col_headers)
        tmp_df.to_excel(out_file, sheet_name="Sheet1", index=False)
        print("Created", out_file)
    try:
        writer = pd.ExcelWriter(
            out_file, engine="openpyxl", mode="a", if_sheet_exists="overlay"
        )
        df.to_excel(
            writer,
            sheet_name="Sheet1",
            startrow=writer.sheets["Sheet1"].max_row,
            header=False,
            index=False,
        )
        writer.save()
        print(f"===== Updated {out_file}. {len(df)} row(s) added to excel file =====")
    except Exception as Err:
        print("ERROR WRITING XLSX", Err)


## update rows after it is done
def update_cell(nbs, status):
    for nb in nbs:
        wb = openpyxl.load_workbook("session.xlsx")
        wb["Sheet1"][f"E{nb}"] = status
        wb.save("session.xlsx")


## start_stop
def start_stop():
    wb = openpyxl.load_workbook("input.xlsx")
    sheet1 = wb.active
    dates = []
    for i in range(1, sheet1.max_row + 1):
        if i > 5:
            break
        temp = sheet1.cell(row=i, column=1).value
        Year = sheet1.cell(row=i, column=2).value
        Month = sheet1.cell(row=i, column=3).value
        Day = sheet1.cell(row=i, column=4).value
        if temp:
            temp = temp.strip().lower()
        if temp == "start date" or temp == "end date":
            dates.append(Day + "." + Month + "." + Year)
    start = dates[0]
    stop = dates[-1]
    return start, stop


## create new session
def create_session():
    wb = openpyxl.load_workbook("input.xlsx")
    sheet1 = wb.active
    data = []
    for i in range(1, sheet1.max_row + 1):
        js = {}
        js["Country"] = sheet1.cell(row=i, column=1).value
        js["League"] = sheet1.cell(row=i, column=2).value
        js["Conference"] = sheet1.cell(row=i, column=3).value
        js["Team"] = sheet1.cell(row=i, column=4).value
        js["Status"] = sheet1.cell(row=i, column=5).value
        if not js["Status"] == "Yes":
            continue
        data.append(js)
    df = pd.DataFrame.from_records(data)
    df.to_excel("session.xlsx", sheet_name="Sheet1", index=False)
    print("New session created")


## read session file xlsx
def read_session():
    wb = openpyxl.load_workbook("session.xlsx")
    sheet1 = wb.active
    data = []
    for i in range(1, sheet1.max_row + 1):
        Country = sheet1.cell(row=i, column=1).value
        League = sheet1.cell(row=i, column=2).value
        Conference = sheet1.cell(row=i, column=3).value
        Team = sheet1.cell(row=i, column=4).value
        Select = sheet1.cell(row=i, column=5).value
        if not Select == "Yes":
            continue
        data.append((i, Country, League, Conference, Team, Select))
    if len(data) == 0:
        sys.exit("Session completed. Delete session.xlsx file to start new session.")
        return None
    (
        start,
        stop,
    ) = start_stop()
    return start, stop, data


################################################ main #######################################################

if os.path.exists("session.xlsx"):
    start, stop, teams = read_session()
    print("Continuing existing session. . .")
else:
    create_session()
    start, stop, teams = read_session()
dfs = []
nbs = []
counter = 0
driver = login()
for i, x in enumerate(teams, 1):
    try:
        if i > 1:
            driver.find_element(By.CLASS_NAME, "ae-logo").click()
    except Exception as Err:
        if driver:
            driver.quit()
        print("XXX Possibly logged out. Trying to login again! XXX")
        driver = login()

    try:
        index, country, league, conference, team, _ = x
        print(f"[{i}/{len(teams)}] row-[{index}] {country} {league} {team}")

        driver, _ = click_items(driver, x)
        if not _:
            continue

        table = get_soup(index, driver)
        if not table:
            continue

        js_data = xtract_data(start, stop, table)
        df = pd.DataFrame.from_records(js_data)
        if df.empty:
            update_cell([index], "Null")
            continue
        dfs.append(df)
        nbs.append(index)
        print(f"{len(df)} - rows found. To write {len(dfs)}/{write_frequency}")
        print("-" * 70)

        if len(dfs) == write_frequency:
            data_frame = pd.concat(dfs, ignore_index=True)
            write_excel(data_frame)
            update_cell(nbs, "Done")
            table = None
            dfs = []
            data_frame = None
            print("-" * 70)
        counter += 1
    except Exception as Err:
        print("ERROR: ", Err)
        print("x" * 70)

if dfs:
    data_frame = pd.concat(dfs, ignore_index=True)
    write_excel(data_frame)
    update_cell(nbs, "Done")
    print("#" * 70)
logout(driver)
